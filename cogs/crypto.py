import discord
from discord.ext import commands, tasks
from openpyxl import load_workbook
import urllib.request
import re
import time
from datetime import date


uid = '465475553458388994'

h1 = '0605'
h2 = '1430'
h3 = '2315'

c1 = '/bitcoin/'
c2 = '/ethereum/'
c3 = '/binance-coin/'
c4 = '/solana/'
c5 = '/cardano/'


class crypto(commands.Cog):

    def __init__(self, bot):
        self.bot = bot
        self.cl.start()
        self.mamial.start()


    @commands.command()
    async def cp(self, ctx):
        webdata1 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c1)
        cdata1 = re.findall(r"<span>\$(\S{9})", webdata1.read().decode())
        webdata2 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c2)
        cdata2 = re.findall(r"<span>\$(\S{8})", webdata2.read().decode())
        webdata3 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c2)
        cdata3 = re.findall(r"<span>\$(\S{6})", webdata3.read().decode())
        webdata4 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c4)
        cdata4 = re.findall(r"<span>\$(\S{6})", webdata4.read().decode())
        webdata5 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c5)
        cdata5 = re.findall(r"<span>\$(\S{4})", webdata5.read().decode())
        embed = discord.Embed(title="", description=f"")
        embed.add_field(name='BTC', value=cdata1[0], inline=False)
        embed.add_field(name="ETH", value=cdata2[0], inline=False)
        embed.add_field(name="BNB", value=cdata3[0], inline=False)
        embed.add_field(name="SOL", value=cdata4[0], inline=False)
        embed.add_field(name="ADA", value=cdata5[0], inline=False)
        await ctx.send(embed=embed)

    @tasks.loop(seconds=60)
    async def cl(self):
        await self.bot.wait_until_ready()
        webdata1 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c1)
        cdata1 = re.findall(r"<span>\$(\S{9})", webdata1.read().decode())
        webdata2 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c2)
        cdata2 = re.findall(r"<span>\$(\S{8})", webdata2.read().decode())
        webdata3 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c2)
        cdata3 = re.findall(r"<span>\$(\S{6})", webdata3.read().decode())
        webdata4 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c4)
        cdata4 = re.findall(r"<span>\$(\S{6})", webdata4.read().decode())
        webdata5 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c5)
        cdata5 = re.findall(r"<span>\$(\S{4})", webdata5.read().decode())
        embed = discord.Embed(title="Cryptocurrency Values", description=f"")
        embed.add_field(name="BTC", value=cdata1[0], inline=False)
        embed.add_field(name="ETH", value=cdata2[0], inline=False)
        embed.add_field(name="BNB", value=cdata3[0], inline=False)
        embed.add_field(name="SOL", value=cdata4[0], inline=False)
        embed.add_field(name="ADA", value=cdata5[0], inline=False)
        channel = self.bot.get_channel(id=927667259495829555)
        myid = '<@' + uid + '>'
        if time.strftime("%H%M") == h1:
            await channel.send(myid)
            await channel.send(embed=embed)
        elif time.strftime("%H%M") == h2:
            await channel.send(myid)
            await channel.send(embed=embed)
        elif time.strftime("%H%M") == h3:
            await channel.send(myid)
            await channel.send(embed=embed)

    @tasks.loop(seconds=60)
    async def mamial(self):
        await self.bot.wait_until_ready()
        webdata1 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + c1)
        cdata1 = re.findall(r"<span>\$(\S{2})", webdata1.read().decode())
        embed = discord.Embed(title="", description=f"")
        embed.add_field(name="BTC", value=cdata1[0], inline=False)
        channel = self.bot.get_channel(id=927667259495829555)
        myid = '<@' + uid + '>'
        xl = load_workbook('./cogs/db/crypto.xlsx')
        edit = xl['Sheet1']
        b1 = str(edit['b1'].value)
        b2 = str(edit['b2'].value)
        a1 = str(edit['a1'].value)
        a2 = str(edit['a2'].value)
        if cdata1[0] <= a1[5:] and '1' in b1:
            await channel.send(myid)
            await channel.send(embed=embed)
            edit['b1'].value = 'nmp'
            xl.save('./cogs/db/crypto.xlsx')
        elif cdata1[0] >= a2[5:] and '1' in b2:
            await channel.send(myid)
            await channel.send(embed=embed)
            edit['b2'].value = 'nmp'
            xl.save('./cogs/db/crypto.xlsx')

    @commands.command()
    async def rs(self, ctx):
        xl = load_workbook('./cogs/db/crypto.xlsx')
        edit = xl['Sheet1']
        edit['b1'].value = '1'
        edit['b2'].value = '1'
        a1 = str(edit['a1'].value)
        a2 = str(edit['a2'].value)
        xl.save('./cogs/db/crypto.xlsx')
        await ctx.send('ok ' + a1 + ' ' + a2)

    @commands.command()
    async def nmin(self, ctx):
        xl = load_workbook('./cogs/db/crypto.xlsx')
        edit = xl['Sheet1']
        edit['a1'].value = ctx.message.content
        a1 = str(edit['a1'].value)
        edit['b1'].value = '1'
        xl.save('./cogs/db/crypto.xlsx')
        await ctx.send('nmin ok ' + a1)

    @commands.command()
    async def nmax(self, ctx):
        xl = load_workbook('./cogs/db/crypto.xlsx')
        edit = xl['Sheet1']
        edit['a2'].value = ctx.message.content
        a2 = str(edit['a2'].value)
        edit['b2'].value = '1'
        xl.save('./cogs/db/crypto.xlsx')
        await ctx.send('nmax ok ' + a2)

    @commands.command()
    async def fear(self, ctx):
        today = date.today()
        y = today.strftime("%Y")
        m = today.strftime("%m")
        d = today.strftime("%d")
        await ctx.send(
            'https://alternative.me/images/fng/crypto-fear-and-greed-index-' + y + '-' + m[1:] + '-' + d + '.png')


def setup(bot):
    bot.add_cog(crypto(bot))
