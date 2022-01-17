import os
import discord
from dotenv import load_dotenv
from discord.ext import tasks, commands
from openpyxl import load_workbook
from random import randint, choice
import urllib.request
from googleapiclient.discovery import build
import re


key=''
ytid = 'UCI7ktPB6toqucpkkCiolwLg'
playlistid = '&list=PL8mPWv3h4qJcxyNXAINRXEqHF06cc1Euq'


class yt(commands.Cog):

    def __init__(self, bot):
        self.bot = bot
        self.panvideoalert.start()


    @commands.command()
    async def h(self, ctx):
        embed = discord.Embed(title="PAN COMMANDS", description=f"Here is the list of PAN commands:")
        embed.add_field(name="FAV", value='Random video from a list of the best Pan piano videos', inline=False)
        embed.add_field(name="PIANO", value='Random video of the Pan piano channel', inline=False)
        embed.add_field(name="NEW", value='Spam the new Pan piano video', inline=False)
        await ctx.send(embed=embed)

    @commands.command()
    async def fav(self, ctx):
        xl = load_workbook('./cogs/db/ppyl.xlsx')
        edit = xl['Sheet1']
        r = [edit.cell(row=i, column=1).value for i in range(1, 15)]
        rd = choice(r)
        await ctx.send(rd)

    @commands.command()
    async def piano(self, ctx):
        xl = load_workbook('./cogs/db/ppyl.xlsx')
        edit = xl['Sheet1']
        c1 = str(edit['b1'].value)
        playlist = urllib.request.urlopen('https://www.youtube.com/watch?v=' + c1 + playlistid)
        video_idsplus = re.findall(r"watch\?v=(\S{11})", playlist.read().decode())
        rd = choice(video_idsplus)
        r = ("https://www.youtube.com/watch?v=" + rd)
        yt = build('youtube', 'v3', developerKey=key)
        request = yt.videos().list(
            part='snippet',
            id=rd
        )
        resp = request.execute()
        if re.search(ytid, str(resp)) is not None:
            await ctx.send(r)
        else:
            await self.piano(self, ctx)
            print('funca')

    @commands.command()
    async def new(self, ctx):
        xl = load_workbook('./cogs/db/ppyl.xlsx')
        edit = xl['Sheet1']
        c1 = str(edit['b1'].value)
        await ctx.send("https://www.youtube.com/watch?v=" + c1)
        await ctx.send("https://www.youtube.com/watch?v=" + c1)
        await ctx.send("https://www.youtube.com/watch?v=" + c1)
        await ctx.send("https://www.youtube.com/watch?v=" + c1)

    @tasks.loop(seconds=60)
    async def panvideoalert(self):
        await self.bot.wait_until_ready()
        xl = load_workbook('./cogs/db/ppyl.xlsx')
        edit = xl['Sheet1']
        ytvideos = urllib.request.urlopen('https://www.youtube.com/channel/' + ytid + '/videos')
        video_ids = re.findall(r"watch\?v=(\S{11})", ytvideos.read().decode())
        edit.cell(row=1, column=2).value = video_ids[0]
        c1 = str(edit['b1'].value)
        c2 = str(edit['b2'].value)
        channel = self.bot.get_channel(id=932689791538036806)
        id = '@everyone'
        if c1 in c2:
            xl.save('./cogs/db/ppyl.xlsx')
        else:
            edit['b2'].value = c1
            print("https://www.youtube.com/watch?v=" + c1)
            xl.save('./cogs/db/ppyl.xlsx')
            embed = discord.Embed(title="!!!PAN PIANO NEW VIDEO!!!",
                                  description=f"Pan Piano just uploaded a new video to his channel")
            await channel.send(id)
            await channel.send(embed=embed)
            await channel.send("https://www.youtube.com/watch?v=" + c1)



def setup(bot):
    bot.add_cog(yt(bot))