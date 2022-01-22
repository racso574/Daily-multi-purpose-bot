import discord
from discord.ext import commands
from googletrans import Translator
from openpyxl import load_workbook
import time
from random import randint, choice




class interpreter(commands.Cog):

    def __init__(self, bot):
        self.bot = bot


    @commands.command()
    async def tr(self, ctx, reaction,):
        tr = Translator()
        trf = tr.translate(ctx.message.content[3:], dest='es')
        ms = await ctx.send(ctx.message.content[3:] + ' = ' + trf.text)
        await ms.add_reaction('💾')

    @commands.Cog.listener()
    async def on_raw_reaction_add(self, payload):
        channel = await self.bot.fetch_channel(payload.channel_id)
        message = await channel.fetch_message(payload.message_id)
        user = await self.bot.fetch_user(payload.user_id)
        emoji = payload.emoji
        xl = load_workbook('./cogs/db/ing.xlsx')
        edit = xl['Sheet1']
        c1 = int(edit['b1'].value)
        r = str([edit.cell(row=i, column=1).value for i in range(1, 1000)])
        if channel.id == 927666821761495133 and user.id != 864541492281868320 and message.content not in r:
            edit.cell(row=int(c1), column=1).value = str(message.content)
            await message.add_reaction('✅')
            edit['b1'].value = c1 + 1
            xl.save('./cogs/db/ing.xlsx')
            print(r)



    @commands.command()
    async def hts(self, ctx):
        c = str(ctx.message.content[4:].replace(ctx.message.content[6:], ''))
        tr = Translator()
        trf = tr.translate(str(ctx.message.content[6:]), dest=str(c))
        await ctx.send(ctx.message.content[6:] + ' = ' + trf.text)

    @commands.command()
    async def lg(self, ctx):
        embed = discord.Embed(title="LANGUAGES", description=f"")
        embed.add_field(name="ca", value='catalan', inline=False)
        embed.add_field(name="es", value='spanish', inline=False)
        embed.add_field(name="en", value='english', inline=False)
        embed.add_field(name="ja", value='japanese', inline=False)
        embed.add_field(name="ko", value='korean', inline=False)
        embed.add_field(name="ro", value='romanian', inline=False)
        embed.add_field(name="it", value='italian', inline=False)
        embed.add_field(name="de", value='german', inline=False)
        embed.add_field(name="fr", value='french', inline=False)
        embed.add_field(name="ru", value='russian', inline=False)
        embed.add_field(name="la", value='latin', inline=False)
        embed.add_field(name="pt", value='portuguese', inline=False)
        await ctx.send(embed=embed)


    @commands.command()
    async def wl(self, ctx):
        xl = load_workbook('./cogs/db/ing.xlsx')
        edit = xl['Sheet1']
        v1 = int(edit['b1'].value)
        v1f = v1 - 1
        v2 = int(edit['b2'].value)
        vx = str([edit.cell(row=v2, column=1).value ])
        if v1f != v2:
            print(vx, 'hola')
            await ctx.send(vx)
            edit['b2'].value = v2 + 1
            xl.save('./cogs/db/ing.xlsx')
            await self.wl(self, ctx)
        else:
            edit['b2'].value = 1
            xl.save('./cogs/db/ing.xlsx')






def setup(bot):
    bot.add_cog(interpreter(bot))



