import discord
from discord.ext import commands
from googletrans import Translator
from openpyxl import load_workbook
import time
from random import randint, choice




class misiones(commands.Cog):

    def __init__(self, bot):
        self.bot = bot

    @commands.command()
    async def ms(self, ctx):
        xl = load_workbook('./cogs/db/misiones.xlsx')
        edit = xl['Sheet1']

        edit['b1'].value = 1
        edit['b2'].value = 2
        edit['b3'].value = 3
        edit['b4'].value = 4
        edit['b5'].value = 5
        xl.save('./cogs/db/misiones.xlsx')

        n1 = int('1')
        n2 = int('2')
        n3 = int('3')
        n4 = int('4')
        n5 = int('5')

        i1 = edit.cell(row=n1, column=1).value
        i2 = edit.cell(row=n2, column=1).value
        i3 = edit.cell(row=n3, column=1).value
        i4 = edit.cell(row=n4, column=1).value
        i5 = edit.cell(row=n5, column=1).value

        embed = discord.Embed(title="MISIONES", description=f"")
        embed.add_field(name=n1, value=i1, inline=False)
        embed.add_field(name=n2, value=i2, inline=False)
        embed.add_field(name=n3, value=i3, inline=False)
        embed.add_field(name=n4, value=i4, inline=False)
        embed.add_field(name=n5, value=i5, inline=False)
        ms = await ctx.send(embed=embed)
        await ms.add_reaction('◀')
        await ms.add_reaction('▶')


    @commands.Cog.listener()
    async def on_raw_reaction_add(self, payload):
        channel = await self.bot.fetch_channel(payload.channel_id)
        message = await channel.fetch_message(payload.message_id)
        user = await self.bot.fetch_user(payload.user_id)
        emoji = payload.emoji
        xl = load_workbook('./cogs/db/misiones.xlsx')
        edit = xl['Sheet1']




        if channel.id == 935238981887619123 and user.id != 864541492281868320:
            print(emoji)

            n1 = int(edit['b1'].value) + 5
            n2 = int(edit['b2'].value) + 5
            n3 = int(edit['b3'].value) + 5
            n4 = int(edit['b4'].value) + 5
            n5 = int(edit['b5'].value) + 5
            xl.save('./cogs/db/misiones.xlsx')
            xl = load_workbook('./cogs/db/misiones.xlsx')
            edit = xl['Sheet1']

            i1 = str([edit.cell(row=n1, column=1).value])
            i2 = str([edit.cell(row=n2, column=1).value])
            i3 = str([edit.cell(row=n3, column=1).value])
            i4 = str([edit.cell(row=n4, column=1).value])
            i5 = str([edit.cell(row=n5, column=1).value])



            embed = discord.Embed(title="MISIONES", description=f"")
            embed.add_field(name=n1, value=i1, inline=False)
            embed.add_field(name=n2, value=i2, inline=False)
            embed.add_field(name=n3, value=i3, inline=False)
            embed.add_field(name=n4, value=i4, inline=False)
            embed.add_field(name=n5, value=i5, inline=False)

            await message.edit(embed=embed)



    @commands.command()
    async def nms(self, ctx):
        xl = load_workbook('./cogs/db/misiones.xlsx')
        edit = xl['Sheet1']

        t1 = str(edit['a1'].value)
        t2 = str(edit['a2'].value)
        t3 = str(edit['a3'].value)
        t4 = str(edit['a4'].value)
        t5 = str(edit['a4'].value)

        if t1 == 'None':
            edit['a1'].value = ctx.message.content[4:]
            xl.save('./cogs/db/misiones.xlsx')
        elif t2 == 'None':
            edit['a2'].value = ctx.message.content[4:]
            xl.save('./cogs/db/misiones.xlsx')
        elif t3 == 'None':
            edit['a3'].value = ctx.message.content[4:]
            xl.save('./cogs/db/misiones.xlsx')
        elif t4 == 'None':
            edit['a4'].value = ctx.message.content[4:]
            xl.save('./cogs/db/misiones.xlsx')
        elif t5 == 'None':
            edit['a5'].value = ctx.message.content[4:]
            xl.save('./cogs/db/misiones.xlsx')








def setup(bot):
    bot.add_cog(misiones(bot))
